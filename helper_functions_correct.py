import pandas as pd
from openpyxl import load_workbook
import string
import numpy as np
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import subprocess
import json
import datetime as dt
import yaml
import os

# from pydrive.auth import GoogleAuth
# from pydrive.drive import GoogleDrive
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
from sql_queries import is_replenishment_empty


def write_to_csv(query_df, output_file_name):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :return:
    '''

    # OUTPUT_LOCAL_FILENAME = '/ldap_home/shopeebi/workspace/business_analytics/Salesforce_validation/validation_%s_%s.csv' % (
    #     country, Today)
    username = os.getcwd().split('/')[-1]
    OUTPUT_FILENAME = 'user/{u}/tmp/atp_temp.csv'.format(u=username)
    print(OUTPUT_FILENAME)
    reqHeaders = query_df.columns
    local_filename = output_file_name
    print(local_filename)
    query_df.write.format('csv').mode('overwrite').options(header='false', escape='"', encoding='UTF-8').save(
        OUTPUT_FILENAME)

    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        OUTPUT_FILENAME, local_filename + "_tmp"),
                    shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rmr %s' % OUTPUT_FILENAME, shell=True)
    dbinputfile = open(local_filename + "_tmp", 'rb')
    dboutputfile = open(local_filename, 'wb')
    dboutputfile.write(','.join(reqHeaders) + '\n')

    for line in dbinputfile:
        dboutputfile.write(line)

    dboutputfile.close()

    os.remove(local_filename + "_tmp")


def append_df_to_excel(template_file_name, write_dict, output_file_name, start_row=0, start_col=0, header_bool=True,
                       **to_excel_kwargs):
    '''
    This function takes in the template file name to refer to and a dictionary of dataframes. It then writes and saves to an specified output file.
    :param template_file_name: File name of the template to write to
    :type template_file_name: str
    :param write_dict: Dictionary of Dataframes to write
    :type write_dict: dict()
    :param output_file_name: Output File Name to Save to
    :type output_file_name: str
    :param start_row: Start Row to Write Dataframes from
    :type start_row: int
    :param start_col: Start column to write Dataframes from
    :type start_col: int
    :param header_bool: Boolean whether to write headers or not. True means write headers
    :type header_bool: boolean
    :param to_excel_kwargs: Other excel keyword arguments
    :return:
    '''
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(template_file_name, engine='openpyxl')
    print(writer)
    print(type(writer))
    # writer = pd.ExcelWriter(template_file_name, engine='xlsxwriter')
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    print("Reading Template... {}".format(template_file_name))
    writer.book = load_workbook(template_file_name)
    try:
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    print("Template Read!")
    for sheet_name, value in write_dict.items():
        # Write Date for these Worksheets
        sheet_names_to_write_date = [
            # "SKU details - ATP - W-4",
            "SKU details - ATP - W-2",
            "SKU details - ATP - Yesterday",
            "SKU detail - Replenishment",
            "SKUs - Stock Sync issue",
            "SKUs - OOS 14 days or less",
            "SKUs - PR but no PO",
            'SKUs - OOS more than 14 days',
            'SKUs - PO but no inbound'
        ]
        if sheet_name in sheet_names_to_write_date:
            # value for these sheet_names are different as these are stored in a dictionary
            df = value['df']
            # write out the new sheet
            print("Writing for sheet: {}...".format(sheet_name))
            df.to_excel(writer, sheet_name, index=False, header=header_bool, startrow=start_row, startcol=start_col)
            print("Written to template for sheet: {}!".format(sheet_name))
            ws = writer.book[sheet_name]
            date_row = start_row - 1
            date_col = start_col + 2
            date_index = get_column_letter(date_col) + str(date_row)
            if sheet_name != "SKU detail - Replenishment":
                date = value['date'].strftime("%Y-%m-%d")
                ws[date_index] = date
            else:
                dates = value['dates']
                yesterday_date_index = date_index
                # dates is a dictionary where four_week_ago and yesterday are keys containing values for dates
                four_weeks_ago_date = dates['four_weeks_ago'].strftime("%Y-%m-%d")
                yesterday_date = dates['yesterday'].strftime("%Y-%m-%d")
                ws[yesterday_date_index] = yesterday_date
                # write for four weeks ago column TODO: Make this value not hardcodded
                four_weeks_date_row = start_row - 1
                four_weeks_date_col = start_col + 4
                four_weeks_date_index = get_column_letter(four_weeks_date_col) + str(four_weeks_date_row)
                ws[four_weeks_date_index] = four_weeks_ago_date
        else:
            # write out the new sheet
            print("Writing for sheet: {}...".format(sheet_name))
            value.to_excel(writer, sheet_name, index=False, header=header_bool, startrow=start_row, startcol=start_col)
            print("Written to template for sheet: {}!".format(sheet_name))
    # save the workbook
    print("Saving Workbook...")
    writer.book.save(output_file_name)
    print("Completed!")


def regional_atp_rewrite(ATP_path, template_path, c):
    '''
    This function takes in the template file to refer to and a dictionary of dataframes. It then writes and saves to an specified output file.
    :param template_path: File path of the template for ATP summary report to write to
    :type template_path: str
    :param ATP_path: File path of the ATP report to get the summary data from
    :type ATP_path: str
    :param c: country sequese in the country list
    :type c: int
    :return:
    '''
    writer = pd.ExcelWriter(ATP_path, engine='openpyxl')
    writer.book = load_workbook(ATP_path, data_only=True, read_only=True)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    writer2 = pd.ExcelWriter(template_path, engine='openpyxl')
    writer2.book = load_workbook(template_path)
    writer2.sheets = {ws.title: ws for ws in writer.book.worksheets}
    ws = writer.book['Summary']
    ws_temp = writer2.book['Summary']
    n = ['H', 'L', 'AB']
    ls1 = [8, 9, 16, 17, 18, 19, 20]
    ls2 = [7, 8, 10, 11, 12, 13, 14]
    for j in range(3):
        v = chr(ord('G') + 8 * j + c)
        if ord(v) > 90:  # chr(90)=Z
            v = 'A' + chr((ord(v) - 90) + 64)
        for i in range(len(ls1)):
            c1 = '%s%d' % (n[j], ls1[i])  # loc cell
            c2 = '%s%d' % (v, ls2[i])
            cell = ws[c1].value
            ws_temp[c2].value = cell
        others = 0
        for i in range(10, 16):
            o1 = '%s%d' % (n[j], i)  # loc cell
            others += ws[o1].value
        o2 = '%s%d' % (v, 9)
        ws_temp[o2].value = others
    try:
        cell = ws['H5'].value
        ws_temp['C3'].value = 'Date: ' + cell
    except:
        print('cannot concatenate str and NoneType objects')
    writer2.book.save(template_path)
    return


def regional_replenishment_rewrite(Rep_path, template_path, c):
    '''
    This function takes in the template file to refer to and a dictionary of dataframes. It then writes and saves to an specified output file.
    :param template_path: File path of the template for replenishment summary report to write to
    :type template_path: str
    :param Rep_path: File path of the Replenishment report to get the summary data from
    :type Rep_path: str
    :param c: country sequese in the country list
    :type c: int
    :return:
    '''
    wb = load_workbook(Rep_path, data_only=True, read_only=True)
    Rep_template = load_workbook(template_path)
    ws = wb['Summary - SKUs hit ROP']
    ws_temp = Rep_template['Summary']
    n = 'J'
    ls = [(11, 9), (12, 10), (13, 11), (14, 12), (15, 13), (16, 14), (17, 15), (18, 16), (19, 17), (22, 18), (23, 19),
          (24, 20)]
    v = chr(ord('J') - 1 + c)
    for i in ls:
        c1 = '%s%d' % (n, i[0])  # loc cell
        c2 = '%s%d' % (v, i[1])
        cell = ws[c1].value
        ws_temp[c2].value = cell
    try:
        cell = ws['D7'].value
        ws_temp['C6'].value = cell
    except:
        print('cannot concatenate str and NoneType objects')
    Rep_template.save(template_path)
    return


def load_report_path(filepath):
    fh = open(filepath)
    data = json.load(fh)
    fh.close()
    return (data)


def save_report_path(filename, data):
    with open(filename, 'w')as f:
        json.dump(data, f)
