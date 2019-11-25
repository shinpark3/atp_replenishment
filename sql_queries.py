# -*- coding: utf-8 -*-
from __future__ import division
import csv
import sys

# reload(sys)
# sys.setdefaultencoding('utf-8')
from collections import defaultdict
from pyspark import SparkContext, SparkConf
from pyspark.sql import SQLContext, HiveContext, SparkSession
from pyspark.sql.column import Column
from pyspark.sql.dataframe import DataFrame
from pyspark.sql.functions import udf, when, lit, asc, broadcast
from pyspark.sql import Row
from dateutil.rrule import rrule, DAILY
import subprocess
import os
import time
import math
import operator
import collections
import zipfile

sys.path.append("/opt/noattach")
# from noattach import NoAttach
# import sendmail
from os.path import dirname, basename
import glob
from dateutil.relativedelta import relativedelta
from decimal import *
from pyspark.sql.types import *
import pandas as pd
import urllib
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import ast
import json
from decimal import *
from openpyxl.styles import Alignment
import base64
import pyspark
# from bi_address_book import send_mail
import re
import os
import datetime
import time
import calendar

ZIP_SIZE_LIMIT = '800M'


SPARK_CONF = SparkConf().set('spark.locality.wait', '1000ms').set('spark.executor.memory', '8G').set(
    'spark.cores.max', '100')
name = 'replenishment_atp'
sc = SparkContext(appName=name, conf=SPARK_CONF)
spark = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
spark.sql('use shopee')


def is_replenishment_empty(date, country):
    '''
    Checks if replenishment table is empty for that specific date. If it is empty, return True
    :param date: lookback date start value
    :type date: Datetime
    :return is_empty: boolean value of whether date is empty
    :type is_empty: boolean
    '''
    date = date.strftime("%Y-%m-%d")
    query = '''
    SELECT
        sku_id, grass_date
    FROM 
        shopee_bi_replenishment_data_pool_core a
    INNER JOIN
        (SELECT sku_id as c_sku_id, grass_date as c_grass_date 
        from 
            shopee_bi_sbs_mart 
        where 
            grass_region = '{region}' and grass_date = date('{check_date}')
        ) as c
    ON 
        a.sku_id = c.c_sku_id
    AND
        a.grass_date = c.c_grass_date
    WHERE 
        a.grass_date = date('{check_date}')
    AND 
        a.grass_region = '{region}'
    AND 
        a.contract_type in ('B2C', 'B2B2C') 
    LIMIT 1
    '''.format(check_date=date, region=country)

    # if one row of data is returned, replenishment table isnt empty
    is_empty = spark.sql(query).toPandas().shape[0] != 1
    return is_empty


def get_latest_r_date(country):
    print("Getting Latest Date from Replenishment Core")
    '''
    Get latest date from replenishment_data_pool_core so that dates correspond with one another.
    :param country:
    :return:
    '''
    query = '''
    select 
        MAX(grass_date) as latest_date 
    FROM
        shopee_bi_replenishment_data_pool_core
    WHERE
        grass_region = '{}'
    '''.format(country)
    print('starting')
    date = spark.sql(query).toPandas()
    print('done')
    return date.latest_date.astype('datetime64').max()


def get_brands(country, four_weeks_ago, two_weeks_ago, yesterday):
    print("Getting Brands ...")

    query = ''' 
        select *
        from 
        (select a.sku_id, b.* from
            (SELECT DISTINCT(sku_id) FROM shopee_bi_replenishment_data_pool_core where grass_date in ('{}', '{}', '{}')) a
            LEFT JOIN 
                (SELECT 
                    *
                FROM
                    shopee_bi_brand_info
                WHERE
                    grass_region = '{}'
                ) b
            ON 
                cast(split(a.sku_id,'_')[0] as BIGINT) = b.itemid) k 
        WHERE k.grass_region = '{}'
        '''.format(yesterday, two_weeks_ago, four_weeks_ago, country, country)

    return spark.sql(query)


def get_cogs_color(country, r_date):
    print("Getting Cogs")
    start_date = (r_date - datetime.timedelta(days=7*13)).strftime('%Y-%m-%d')
    query = ''' 
    select sku_id, grass_date, cogs, color, stock_on_hand from shopee_bi_sbs_mart where grass_region = '{}' and grass_date >= date('{}')
    '''.format(country, start_date)

    return spark.sql(query)


def get_color(country, r_date):
    print("Getting Color...")
    start_date = (r_date - datetime.timedelta(days=7*13)).strftime('%Y-%m-%d')
    query = ''' select sku_id, color, stock_on_hand, grass_date from shopee_bi_sbs_mart where grass_region = '{}' and grass_date >= date('{}')
    '''.format(country, start_date)

    return spark.sql(query)


def get_cat_map(country):
    print("Getting Category Map...")
    query = ''' 
    select * from shopee_bi_reg_cat_map where country = '{}'
    '''.format(country)
    return spark.sql(query)


def get_item_profile(country):
    print("Getting Item Profile...")
    query = '''select itemid, modelid, decode(shop_name, 'utf-8') as shop_name, item_price, main_cat from item_profile where grass_region = '{}' and is_b2c = 1
    '''.format(country)

    return spark.sql(query)


def get_pr_po_inbound(country):
    print("Getting pr, po, inbound....")
    query = ''' 
    select 
        pr_sku.sku_id, pr_sku.reason as pr_reason, pr_sku.quantity as pr_quantity, 
        date(from_unixtime(pr_tab.ctime)) as pr_date, pr_tab.country, 
        pr_tab.request_id as pr_id, pr_sku.sourcing_status, 
        pr_tab.status as pr_status, pr_tab.status_text, po_hourly.order_id,
        po_hourly.status as po_status,
        date(from_unixtime(po_hourly.ctime)) as po_date,
        po_sku.confirmed_quantity as po_quantity,
        date(from_unixtime(inbound_tab.inbound_time)) as inbound_time, inbound_tab.quantity as inbound_quantity,
        inbound_tab.inbound_id as inbound_id
    from shopee_pms_db__purchase_request_tab as pr_tab
    left join shopee_pms_db__purchase_request_sku_tab as pr_sku
        on pr_tab.id = pr_sku.pr_id
    left join shopee_pms_db__purchase_order_tab_hourly as po_hourly
        on pr_tab.request_id = po_hourly.request_id
    left join shopee_pms_db__purchase_order_sku_tab as po_sku
        on po_hourly.order_id = po_sku.order_id and pr_sku.sku_id = po_sku.sku_id
    left join shopee_pms_db__purchase_order_inbound_sku_tab as inbound_tab
        on inbound_tab.order_id = po_hourly.order_id and pr_sku.sku_id = inbound_tab.sku_id
    where pr_tab.country = '{}'
    '''.format(country)
    return spark.sql(query)


def get_purchase_type(country):
    print("Getting Purchase Type...")
    query = '''
        SELECT 
        distinct shopee_pms_db__supplier_sku_tab.sku_id as sku_id, 
        shopee_pms_db__supplier_tab.purchase_type as purchase_type
    FROM
        shopee_pms_db__supplier_sku_tab
    LEFT JOIN
        shopee_pms_db__supplier_tab
    ON
        shopee_pms_db__supplier_sku_tab.supplier_id = shopee_pms_db__supplier_tab.supplier_id
    WHERE 
        date(from_unixtime(shopee_pms_db__supplier_sku_tab.ctime)) <= date(current_date)
    AND 
        shopee_pms_db__supplier_tab.country = '{}'
    '''.format(country)
    return spark.sql(query)


def get_cfs(region, today_date, end_date):
    print("Getting CFS ...")

    # table name to query for CFS
    table_name = 'shopee_backend_{}_db__promotion_flash_sale_item_tab'.format(str.lower(region))
    query = """
    select
        concat(itemid, '_', modelid) as sku_id,
        sum(stock) as last_fs_total_stock
    from 
        {region_table}
    where 
        from_unixtime(start_time, 'yyyy-MM-dd') > '{today}'
    group by 
        1        
    """.format(region_table=table_name, today=today_date)
    return spark.sql(query)


def get_replenishment_core(country, r_date):
    print("Getting replenishment core...")
    start_date = (r_date - datetime.timedelta(days=7*13)).strftime('%Y-%m-%d')
    query = '''
                select 
                    * 
                from 
                    shopee_bi_replenishment_data_pool_core a           
                where 
                    a.grass_region = '{cntry}' 
                and 
                    a.contract_type in ('B2C', 'B2B2C') 
                and 
                    a.grass_date >= date('{s_date}') 
    '''.format(cntry=country, s_date=start_date)
    return spark.sql(query)


def get_replenishment_ext(country, four_weeks_ago, two_weeks_ago, yesterday):
    print("Getting Replenishment Ext...")
    query = ''' SELECT * from shopee_bi_replenishment_data_pool_ext_replenishment_info 
    where grass_region = '{}' 
    and grass_date >= '{}' - interval '14' day
    '''.format(country, four_weeks_ago)
    return spark.sql(query)
